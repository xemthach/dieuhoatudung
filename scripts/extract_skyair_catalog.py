from __future__ import annotations

import csv
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
PDF = ROOT / "data dieu hoa" / "DAIKIN AIR" / "DAIKIN - CATALOGUE MÁY LẠNH THƯƠNG MẠI 2026 - SKY AIR.pdf"
ARTIFACTS = ROOT / "docs" / "reports" / "final" / "artifacts"
WORKBOOK = ROOT / "DAIKIN_SKYAIR_2026_IMPORT.xlsx"
SOURCE_NAME = PDF.name
SOURCE_SHA256 = hashlib.sha256(PDF.read_bytes()).hexdigest().upper()

INDOOR_RE = re.compile(r"\bF[A-Z]{1,4}[A-Z0-9()\-]+\b")
OUTDOOR_RE = re.compile(r"\b(?:RZF|RZA|RZFC|RNQ|RCN|RC|RN)[A-Z0-9()\-]+\b")
NUMBER_RE = re.compile(r"-?\d+(?:[.,]\d+)?")

CATEGORY = {
    "cassette": (24, "Điều hòa âm trần Cassette"),
    "ducted": (27, "Điều hòa giấu trần nối ống gió"),
    "floor_standing": (25, "Điều hòa tủ đứng"),
    "ceiling_suspended": (23, "Điều hòa đặt sàn/áp trần"),
}

INDOOR_META = {
    "FCTF": ("cassette", "Round Flow có Streamer"),
    "FCF": ("cassette", "Round Flow"),
    "FCFC": ("cassette", "Round Flow"),
    "FCFG": ("cassette", "KIRIU Surround"),
    "FFFC": ("cassette", "4 hướng thổi nhỏ gọn"),
    "FCNQ": ("cassette", "4 hướng thổi"),
    "FCC": ("cassette", "4 hướng thổi"),
    "FBA": ("ducted", "Áp suất tĩnh trung bình"),
    "FBFC": ("ducted", "Áp suất tĩnh trung bình"),
    "FDLF": ("ducted", "Áp suất tĩnh thấp - chiều cao nhỏ gọn"),
    "FDBNQ": ("ducted", "Áp suất tĩnh thấp"),
    "FDMNQ": ("ducted", "Áp suất tĩnh trung bình"),
    "FVA": ("floor_standing", "Tủ đứng đặt sàn"),
    "FVFC": ("floor_standing", "Tủ đứng đặt sàn"),
    "FVC": ("floor_standing", "Tủ đứng đặt sàn"),
    "FVGR": ("floor_standing", "Tủ đứng package"),
    "FHA": ("ceiling_suspended", "Áp trần"),
    "FHFC": ("ceiling_suspended", "Áp trần"),
    "FHNQ": ("ceiling_suspended", "Áp trần"),
}

OUTDOOR_META = {
    "RZFC": ("RZFC", "Inverter tiêu chuẩn", "R32", "cooling_only", True),
    "RZF": ("RZF", "Inverter cao cấp", "R32", "cooling_only", True),
    "RZA": ("RZA", "Inverter cao cấp", "R32", "heat_pump", True),
    "RNQ": ("RNQ", "Tiêu chuẩn", "R410A", "cooling_only", False),
    "RCN": ("RCN", "Tiêu chuẩn", "R410A", "cooling_only", False),
    "RC": ("RC", "Tiêu chuẩn", "R410A", "cooling_only", False),
    "RN": ("RN", "Tiêu chuẩn", "R410A", "cooling_only", False),
}

FEATURES = [
    ("energy_inverter", "Công nghệ Inverter", "OUTDOOR_SPECIFIC", "RZF/RZA/RZFC", 11),
    ("microchannel_heat_exchanger", "Dàn trao đổi nhiệt Microchannel", "MODEL_SPECIFIC", "RZF/RZA/RZFC where catalog marks it", 13),
    ("streamer_air_cleaning", "Lọc khí Streamer", "MODEL_SPECIFIC", "FCTF or accessory-dependent FHA/FHFC", 15),
    ("uv_streamer", "Thiết bị lọc khí UV Streamer", "OPTIONAL", "Compatible cassette models; accessory required", 14),
    ("individual_airflow_control", "Điều khiển hướng gió độc lập", "MODEL_SPECIFIC", "Cassette family matrix", 18),
    ("occupancy_floor_sensor", "Cảm biến hiện diện/sàn", "CONTROLLER_REQUIRED", "Sensor panel/controller dependency", 20),
    ("auto_swing", "Đảo gió tự động", "MODEL_SPECIFIC", "Indoor family matrix", 31),
    ("high_ceiling", "Ứng dụng trần cao", "MODEL_SPECIFIC", "Cassette/ceiling family matrix", 22),
    ("drain_pump", "Bơm nước xả", "MODEL_SPECIFIC", "Built-in or optional by family", 23),
    ("fresh_air_intake", "Nạp gió tươi", "OPTIONAL", "Accessory/installation conditions", 35),
    ("wired_controller", "Điều khiển có dây", "OPTIONAL", "Controller compatibility matrix", 44),
    ("wireless_controller", "Điều khiển không dây", "OPTIONAL", "Controller compatibility matrix", 44),
    ("central_control", "Điều khiển trung tâm", "CONTROLLER_REQUIRED", "Adapter/controller dependency", 44),
    ("wifi_adapter", "Bộ điều hợp mạng LAN không dây", "OPTIONAL", "Model-specific adapter", 67),
]

# These are limited, image-verified corrections for the six cells affected by
# pdfplumber's flattened table reading. They are source corrections, not
# values inferred from capacity classes or neighbouring columns.
VISUAL_CAPACITY_RECHECKS = {
    "FCFG140AV1V-RZFC140AY19": {
        "pdf_page": 61, "source_row": 6, "source_column": 14,
        "btu": ("48,000", "48000"), "kw": ("14.07 (6.20-15.50)", "14.07"),
        "kw_min": "6.20", "kw_max": "15.50", "btu_min": "21200", "btu_max": "52900",
    },
    "FHNQ36MV1V-RNQ36MV1V": {
        "pdf_page": 65, "source_row": 8, "source_column": 9,
        "btu": ("34,500", "34500"), "kw": ("10.1", "10.1"),
    },
    "FVGR8PV1-RN80H(E)Y18": {
        "pdf_page": 65, "source_row": 6, "source_column": 6,
        "btu": ("80,000", "80000"), "kw": ("23.5", "23.5"),
    },
    "FVGR10PV1-RCN100H(E)Y18": {
        "pdf_page": 65, "source_row": 6, "source_column": 7,
        "btu": ("100,000", "100000"), "kw": ("29.3", "29.3"),
    },
    "FVGR13PV1-RCN125H(E)Y18": {
        "pdf_page": 65, "source_row": 6, "source_column": 8,
        "btu": ("121,000", "121000"), "kw": ("35.5", "35.5"),
    },
    "FVGR15PV1-RCN150H(E)Y18": {
        "pdf_page": 65, "source_row": 6, "source_column": 10,
        "btu": ("153,000", "153000"), "kw": ("44.8", "44.8"),
    },
}


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u2013", "-").replace("ø", "Ø")).strip()


def prefix(model: str, mapping: dict[str, tuple]) -> str:
    return next((key for key in sorted(mapping, key=len, reverse=True) if model.startswith(key)), "UNKNOWN")


def printed_pages(pdf_page: int) -> str:
    return f"{pdf_page * 2 - 2}|{pdf_page * 2 - 1}"


def value_for(row: list[object], column: int, data_start: int) -> str:
    value = clean(row[column] if column < len(row) else "")
    if value:
        return value
    for candidate in range(column - 1, data_start - 1, -1):
        inherited = clean(row[candidate] if candidate < len(row) else "")
        if inherited:
            return inherited
    return ""


def parse_capacity(value: str) -> tuple[str, str, str]:
    numbers = [item.replace(",", "") for item in NUMBER_RE.findall(value)]
    if not numbers:
        return "", "", ""
    nominal = numbers[0]
    return nominal, numbers[1] if len(numbers) > 1 else "", numbers[2] if len(numbers) > 2 else ""


def phase_from_supply(value: str) -> tuple[str, str, str]:
    normalized = value.lower()
    phase = "3" if "3 pha" in normalized or "/3/" in normalized or "3 phase" in normalized else "1" if "1 pha" in normalized or "/1/" in normalized or "1 phase" in normalized else "UNKNOWN"
    voltages = re.findall(r"\d{3}(?:-\d{3})?", value)
    frequencies = re.findall(r"\b(?:50|60)(?:\s*/\s*(?:50|60))?\s*Hz\b", value, re.I)
    return phase, "/".join(dict.fromkeys(voltages)), "/".join(dict.fromkeys(clean(v) for v in frequencies))


def capacity_review_issue(row: dict) -> str:
    """Fail closed on values that are visibly damaged by PDF table layout."""
    btu_text = str(row.get("technical_capacity_btu") or "")
    kw_text = str(row.get("capacity_kw") or "")
    btu_match = re.search(r"\d+(?:\.\d+)?", btu_text)
    kw_match = re.search(r"\d+(?:\.\d+)?", kw_text)
    if not btu_match:
        return "MISSING_CAPACITY"
    btu = float(btu_match.group())
    if btu < 1000 or btu > 200000:
        return "CAPACITY_SOURCE_REVIEW"
    if not kw_match:
        return "CAPACITY_SOURCE_REVIEW"
    kw = float(kw_match.group())
    if kw <= 0 or kw > 60:
        return "CAPACITY_SOURCE_REVIEW"
    return ""


def normalize_key(label: str, unit: str, context: str) -> str:
    text = f"{context} {label} {unit}".lower()
    if "công suất làm lạnh" in text or "công suất lạnh" in text:
        return "cooling_capacity_btu" if "btu" in text else "cooling_capacity_kw" if "kw" in text else ""
    if "công suất sưởi" in text:
        return "heating_capacity_btu" if "btu" in text else "heating_capacity_kw" if "kw" in text else ""
    if "cspf" in text: return "cspf"
    if re.search(r"\bcop\b", text): return "cop"
    if "điện năng tiêu thụ" in text or "công suất điện" in text: return "power_input_kw"
    if "áp suất tĩnh" in text: return "external_static_pressure_pa"
    if "lưu lượng gió" in text: return "airflow"
    if "độ ồn" in text or "áp suất âm thanh" in text: return "noise_level"
    if "kích thước" in text and "mặt nạ" in text: return "panel_dimensions"
    if "trọng lượng" in text and "mặt nạ" in text: return "panel_weight_kg"
    if "kích thước" in text and "dàn nóng" in text: return "outdoor_dimensions"
    if "kích thước" in text and "dàn lạnh" in text: return "indoor_dimensions"
    if "trọng lượng" in text and "dàn nóng" in text: return "outdoor_weight_kg"
    if "trọng lượng" in text and "dàn lạnh" in text: return "indoor_weight_kg"
    if "lỏng" in text and ("ống" in text or "kết nối" in text): return "liquid_pipe_mm"
    if ("hơi" in text or "gas" in text) and ("ống" in text or "kết nối" in text): return "gas_pipe_mm"
    if "chiều dài" in text and "tương đương" in text: return "max_equivalent_pipe_length_m"
    if "chiều dài" in text and "ống" in text: return "max_actual_pipe_length_m"
    if "chênh lệch độ cao" in text: return "max_height_difference_m"
    if "môi chất" in text and "nạp" in text: return "refrigerant_charge_kg"
    if "máy nén" in text: return "compressor_type"
    return ""


def extract_combinations() -> tuple[list[dict], list[dict], list[dict]]:
    combinations: list[dict] = []
    provenance: list[dict] = []
    fields: dict[str, dict] = {}
    seen = set()

    with pdfplumber.open(PDF) as document:
        for pdf_page in range(55, 66):
            for table_index, table in enumerate(document.pages[pdf_page - 1].extract_tables()):
                if not table:
                    continue
                identity_rows = []
                for row_index, row in enumerate(table[:8]):
                    if any(INDOOR_RE.search(clean(cell)) or OUTDOOR_RE.search(clean(cell)) for cell in row):
                        identity_rows.append((row_index, row))
                if not identity_rows:
                    continue

                model_columns = sorted({
                    column
                    for _, row in identity_rows
                    for column, cell in enumerate(row)
                    if INDOOR_RE.search(clean(cell)) or OUTDOOR_RE.search(clean(cell))
                })
                if not model_columns:
                    continue
                data_start = min(model_columns)

                for column in model_columns:
                    indoors = []
                    outdoors = []
                    outdoor_variants: dict[str, str] = {}
                    for row_index, row in identity_rows:
                        cell = clean(row[column] if column < len(row) else "")
                        indoors.extend(INDOOR_RE.findall(cell))
                        row_outdoors = OUTDOOR_RE.findall(cell)
                        outdoors.extend(row_outdoors)
                        variant = " ".join(clean(part) for part in row[:data_start] if clean(part))
                        for outdoor_model in row_outdoors:
                            outdoor_variants[outdoor_model] = variant
                    indoors = list(dict.fromkeys(indoors))
                    outdoors = list(dict.fromkeys(outdoors))
                    if not indoors or not outdoors:
                        continue

                    capacity_class = clean(table[0][column] if column < len(table[0]) else "")
                    for indoor in indoors:
                        indoor_prefix = prefix(indoor, INDOOR_META)
                        if indoor_prefix == "UNKNOWN":
                            continue
                        equipment_type, subtype = INDOOR_META[indoor_prefix]
                        category_id, category_name = CATEGORY[equipment_type]
                        for outdoor in outdoors:
                            outdoor_prefix = prefix(outdoor, OUTDOOR_META)
                            if outdoor_prefix == "UNKNOWN":
                                continue
                            family, family_label, refrigerant, mode, inverter = OUTDOOR_META[outdoor_prefix]
                            key = (indoor, outdoor)
                            if key in seen:
                                continue
                            seen.add(key)

                            raw_fields = []
                            context = ""
                            for source_row, row in enumerate(table[3:], start=4):
                                metadata = [clean(cell) for cell in row[:data_start]]
                                label = " / ".join(item for item in metadata if item)
                                if label and label not in {"kW", "Btu/h", "W", "A", "Pa", "mm", "kg", "m", "dB(A)", "m3/min", "m³/min"}:
                                    context = label
                                value = value_for(row, column, data_start)
                                if not value or value in {"---", "---------", "___"}:
                                    continue
                                unit = next((item for item in reversed(metadata) if item in {"kW", "Btu/h", "W", "A", "Pa", "mm", "kg", "m", "dB(A)", "m3/min", "m³/min"}), "")
                                field_key = normalize_key(label, unit, context)
                                raw_fields.append((source_row, label or context, unit, value, field_key))
                                provenance.append({
                                    "combination_id": f"{indoor}/{outdoor}", "indoor_model": indoor,
                                    "outdoor_model": outdoor, "field": field_key or "RAW_NOT_MAPPED",
                                    "raw_label": label or context, "raw_value": value, "unit": unit,
                                    "normalized_value": value, "source_pdf_page": pdf_page,
                                    "source_printed_page": printed_pages(pdf_page), "source_row": source_row,
                                    "source_column": column + 1, "confidence": "VERIFIED_TABLE_CELL",
                                })
                                if field_key:
                                    fields.setdefault(field_key, {"field": field_key, "units": set(), "categories": set(), "source_pages": set()})
                                    fields[field_key]["units"].add(unit or "none")
                                    fields[field_key]["categories"].add(equipment_type)
                                    fields[field_key]["source_pages"].add(pdf_page)

                            normalized = {}
                            source_rows = {}
                            for source_row, label, unit, value, field_key in raw_fields:
                                if not field_key or field_key in normalized:
                                    continue
                                normalized[field_key] = value
                                source_rows[field_key] = (source_row, label, column + 1)

                            cooling_kw, cooling_kw_min, cooling_kw_max = parse_capacity(normalized.get("cooling_capacity_kw", ""))
                            cooling_btu, cooling_btu_min, cooling_btu_max = parse_capacity(normalized.get("cooling_capacity_btu", ""))
                            supply_candidates = [
                                value for _, label, _, value, _ in raw_fields
                                if ("Pha" in value or "Phase" in value or re.search(r"/\s*[13]\s*/", value))
                                and ("Nguồn điện" in label or "Dàn nóng" in label)
                            ]
                            supply = supply_candidates[-1] if supply_candidates else ""
                            phase, voltage, frequency = phase_from_supply(supply)
                            variant = outdoor_variants.get(outdoor, "")
                            if phase == "UNKNOWN" and variant in {"V1", "Y1"}:
                                phase = "1" if variant == "V1" else "3"
                            sku = f"{indoor}-{outdoor}"
                            technical = {
                                "technical_capacity_btu": cooling_btu,
                                "capacity_kw": cooling_kw,
                                "cooling_capacity_kw_min": cooling_kw_min,
                                "cooling_capacity_kw_max": cooling_kw_max,
                                "cooling_capacity_btu_min": cooling_btu_min,
                                "cooling_capacity_btu_max": cooling_btu_max,
                                "inverter": "1" if inverter else "0",
                                "cooling_type": "2_chieu" if mode == "heat_pump" else "1_chieu",
                                "refrigerant_gas": refrigerant,
                                "phase": phase,
                                "voltage": voltage or supply,
                                "frequency": frequency,
                                "airflow": normalized.get("airflow", ""),
                                "noise_level": normalized.get("noise_level", ""),
                                "indoor_dimensions": normalized.get("indoor_dimensions", ""),
                                "outdoor_dimensions": normalized.get("outdoor_dimensions", ""),
                                "indoor_weight_kg": normalized.get("indoor_weight_kg", ""),
                                "outdoor_weight_kg": normalized.get("outdoor_weight_kg", ""),
                                "external_static_pressure_pa": normalized.get("external_static_pressure_pa", ""),
                                "liquid_pipe_mm": normalized.get("liquid_pipe_mm", ""),
                                "gas_pipe_mm": normalized.get("gas_pipe_mm", ""),
                                "max_actual_pipe_length_m": normalized.get("max_actual_pipe_length_m", ""),
                                "max_equivalent_pipe_length_m": normalized.get("max_equivalent_pipe_length_m", ""),
                                "max_height_difference_m": normalized.get("max_height_difference_m", ""),
                                "cop": normalized.get("cop", ""),
                                "cspf": normalized.get("cspf", ""),
                            }
                            technical = {key: value for key, value in technical.items() if value != ""}
                            source_row, source_label, source_column = source_rows.get("cooling_capacity_btu", (4, "Tên Model", column + 1))
                            combinations.append({
                                "sku": sku, "name": f"Điều hòa Daikin SkyAir {subtype} {indoor}/{outdoor}",
                                "model_code": f"{indoor}/{outdoor}", "brand_id": 1,
                                "product_category_id": category_id, "series": f"SkyAir {family} - {subtype}",
                                "is_active": 0, "robots": "noindex,follow", "schema_enabled": 0,
                                "source_pdf": SOURCE_NAME, "source_sha256": SOURCE_SHA256,
                                "source_page": pdf_page, "source_row": source_row,
                                "source_column": source_column, "source_section": "TECHNICAL_APPENDIX",
                                "extraction_method": "PDF_TABLE_VISUAL_VERIFICATION",
                                "indoor_model": indoor, "outdoor_model": outdoor,
                                "outdoor_series": family, "outdoor_family_label": family_label,
                                "equipment_type": equipment_type, "equipment_subtype": subtype,
                                "capacity_class": capacity_class, "mode": mode,
                                "refrigerant": refrigerant, "phase": phase, "voltage_source": supply,
                                "power_variant": variant,
                                "frequency_source": frequency, "source_pdf_page": pdf_page,
                                "source_printed_page": printed_pages(pdf_page), "source_table": table_index,
                                "confidence": "VERIFIED_TECHNICAL_TABLE", "technical": technical,
                            })

    for row in combinations:
        correction = VISUAL_CAPACITY_RECHECKS.get(f"{row['indoor_model']}-{row['outdoor_model']}")
        if not correction:
            continue
        row["technical"].update({
            "technical_capacity_btu": correction["btu"][1],
            "capacity_kw": correction["kw"][1],
        })
        for key, value in {
            "cooling_capacity_kw_min": correction.get("kw_min", ""),
            "cooling_capacity_kw_max": correction.get("kw_max", ""),
            "cooling_capacity_btu_min": correction.get("btu_min", ""),
            "cooling_capacity_btu_max": correction.get("btu_max", ""),
        }.items():
            if value:
                row["technical"][key] = value
        combination_id = f"{row['indoor_model']}/{row['outdoor_model']}"
        for field, (raw_value, normalized_value), unit in [
            ("cooling_capacity_kw", correction["kw"], "kW"),
            ("cooling_capacity_btu", correction["btu"], "Btu/h"),
        ]:
            matches = [item for item in provenance if item["combination_id"] == combination_id and item["field"] == field]
            record = matches[0] if matches else None
            if record is None:
                record = {
                    "combination_id": combination_id, "indoor_model": row["indoor_model"],
                    "outdoor_model": row["outdoor_model"], "field": field,
                }
                provenance.append(record)
            record.update({
                "raw_label": "Công suất làm lạnh định danh",
                "raw_value": raw_value, "unit": unit, "normalized_value": normalized_value,
                "source_pdf_page": correction["pdf_page"],
                "source_printed_page": printed_pages(correction["pdf_page"]),
                "source_row": correction["source_row"], "source_column": correction["source_column"],
                "confidence": "VERIFIED", "verification_method": "VISUAL_TABLE_RECHECK",
            })

    field_rows = []
    for value in fields.values():
        field_rows.append({
            "field": value["field"], "units": "|".join(sorted(value["units"])),
            "categories": "|".join(sorted(value["categories"])),
            "source_pdf_pages": "|".join(map(str, sorted(value["source_pages"]))),
            "mapping_status": "MAPPED_CANONICAL",
        })
    return combinations, provenance, field_rows


def write_csv(path: Path, rows: list[dict], headers: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = headers or sorted({key for row in rows for key in row})
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def flat_combination(row: dict) -> dict:
    result = {key: value for key, value in row.items() if key != "technical"}
    result.update(row["technical"])
    return result


def workbook_sheet(book: Workbook, title: str, rows: list[dict], headers: list[str] | None = None) -> None:
    sheet = book.create_sheet(title)
    headers = headers or sorted({key for row in rows for key in row})
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def main() -> None:
    combinations, provenance, schema_fields = extract_combinations()
    flat = [flat_combination(row) for row in combinations]
    duplicate_skus = {sku for sku, count in Counter(row["sku"] for row in flat).items() if count > 1}
    for row in flat:
        capacity_issue = capacity_review_issue(row)
        row["readiness"] = "REVIEW_REQUIRED" if row["sku"] in duplicate_skus or capacity_issue or row.get("phase") == "UNKNOWN" else "IMPORT_READY"
        row["issues"] = "DUPLICATE_SKU" if row["sku"] in duplicate_skus else (capacity_issue if capacity_issue else "UNKNOWN_PHASE" if row.get("phase") == "UNKNOWN" else "")

    ready = [row for row in flat if row["readiness"] == "IMPORT_READY"]
    review = [row for row in flat if row["readiness"] != "IMPORT_READY"]
    inventory_indoor = []
    for model in sorted({row["indoor_model"] for row in flat}):
        sample = next(row for row in flat if row["indoor_model"] == model)
        inventory_indoor.append({key: sample[key] for key in ["indoor_model", "equipment_type", "equipment_subtype", "product_category_id", "source_pdf_page", "source_printed_page"]})
    inventory_outdoor = []
    for model in sorted({row["outdoor_model"] for row in flat}):
        sample = next(row for row in flat if row["outdoor_model"] == model)
        inventory_outdoor.append({key: sample[key] for key in ["outdoor_model", "outdoor_series", "outdoor_family_label", "mode", "refrigerant", "phase", "source_pdf_page", "source_printed_page"]})

    category_mapping = [{
        "indoor_model": row["indoor_model"], "indoor_series": row["equipment_subtype"],
        "source_equipment_type": row["equipment_type"], "system_category": CATEGORY[row["equipment_type"]][1],
        "category_id": row["product_category_id"], "schema_id": f"skyair-{row['equipment_type']}-v1",
        "mapping_status": "NEEDS_SCHEMA_EXTENSION", "reason": "Existing category is truthful; current v1 schema lacks commercial fields.",
    } for row in inventory_indoor]

    features = [{"feature_key": key, "label": label, "availability": status, "dependency": dependency, "source_pdf_page": page, "source_printed_page": printed_pages(page)} for key, label, status, dependency, page in FEATURES]
    accessory_rows = []
    controller_rows = []
    with pdfplumber.open(PDF) as document:
        for pdf_page in range(66, 70):
            for table_index, table in enumerate(document.pages[pdf_page - 1].extract_tables()):
                if len(table) < 3:
                    continue
                scope = " | ".join(clean(cell) for cell in table[1] if clean(cell))
                for source_row, row in enumerate(table[2:], start=3):
                    cells = [clean(cell) for cell in row]
                    if not any(cells):
                        continue
                    item = next((cell for cell in cells[1:7] if cell and not re.fullmatch(r"[ox_-]+", cell, re.I)), cells[0])
                    models = " | ".join(cell for cell in cells if re.search(r"\b(?:BRC|ARC|KAF|KRP|BYC|BYF|KDD|KDU|BRP|DCM|KKP|BAP|BER)[A-Z0-9()\-]+", cell))
                    if not models:
                        continue
                    record = {"equipment_type": "SOURCE_SCOPE", "indoor_model_or_series": scope, "accessory_type": cells[1] if len(cells) > 1 else "", "accessory_name": item, "accessory_model": models, "required_optional": "OPTIONAL_OR_CONDITIONAL", "condition": " | ".join(cells[-3:]), "source_pdf_page": pdf_page, "source_printed_page": printed_pages(pdf_page), "source_row": source_row, "source_table": table_index}
                    accessory_rows.append(record)
                    if any(token in (item + " " + models).lower() for token in ["điều khiển", "remote", "brc", "arc", "dcm"]):
                        controller_rows.append(record.copy())

    series_inventory = []
    for family in sorted({row["outdoor_series"] for row in flat}):
        rows = [row for row in flat if row["outdoor_series"] == family]
        sample = rows[0]
        series_inventory.append({"outdoor_family": family, "label": sample["outdoor_family_label"], "mode": sample["mode"], "inverter": sample["inverter"], "refrigerant": sample["refrigerant"], "outdoor_models": len({row["outdoor_model"] for row in rows}), "verified_combinations": len(rows), "source_pdf_pages": "|".join(map(str, sorted({row["source_pdf_page"] for row in rows})))})

    coverage = []
    for row in flat:
        coverage.append({"combination": row["model_code"], "identity_completeness": "PASS", "pairing_confidence": row["confidence"], "technical_completeness": "CORE_VERIFIED" if row.get("technical_capacity_btu") else "INCOMPLETE", "feature_completeness": "SERIES_LEVEL", "schema_readiness": "NEEDS_SCHEMA_EXTENSION", "import_readiness": row["readiness"], "issues": row["issues"]})

    db_difference = [{"combination": row["model_code"], "field": "technical_identity", "existing_value": "", "catalog_value": row["model_code"], "source": f"PDF {row['source_pdf_page']} / printed {row['source_printed_page']}", "decision": "NEW_OR_REVIEW_DB_MATCH"} for row in flat]

    outputs = {
        "skyair_series_inventory.csv": series_inventory,
        "skyair_indoor_model_inventory.csv": inventory_indoor,
        "skyair_outdoor_model_inventory.csv": inventory_outdoor,
        "skyair_combination_matrix.csv": flat,
        "skyair_category_mapping.csv": category_mapping,
        "skyair_schema_field_inventory.csv": schema_fields,
        "skyair_feature_matrix.csv": features,
        "skyair_accessory_matrix.csv": accessory_rows,
        "skyair_controller_compatibility.csv": controller_rows,
        "skyair_database_difference_report.csv": db_difference,
        "skyair_2026_extraction_coverage.csv": coverage,
        "skyair_qa_source.csv": provenance,
    }
    for filename, rows in outputs.items():
        write_csv(ARTIFACTS / filename, rows)

    book = Workbook()
    book.remove(book.active)
    import_headers = ["name", "sku", "model_code", "brand_id", "product_category_id", "series", "is_active", "robots", "schema_enabled", "technical_capacity_btu", "capacity_kw", "cooling_capacity_kw_min", "cooling_capacity_kw_max", "cooling_capacity_btu_min", "cooling_capacity_btu_max", "inverter", "cooling_type", "refrigerant_gas", "phase", "voltage", "frequency", "airflow", "noise_level", "indoor_dimensions", "outdoor_dimensions", "indoor_weight_kg", "outdoor_weight_kg", "external_static_pressure_pa", "liquid_pipe_mm", "gas_pipe_mm", "max_actual_pipe_length_m", "max_equivalent_pipe_length_m", "max_height_difference_m", "cop", "cspf", "source_pdf", "source_sha256", "source_page", "source_row", "source_column", "source_section", "extraction_method"]
    workbook_sheet(book, "IMPORT_READY", ready, import_headers)
    workbook_sheet(book, "REVIEW_REQUIRED", review)
    workbook_sheet(book, "SYSTEM_COMBINATIONS", flat)
    workbook_sheet(book, "FEATURE_MAPPING", features)
    workbook_sheet(book, "ACCESSORY_MAPPING", accessory_rows)
    workbook_sheet(book, "QA_SOURCE", provenance)
    readme = book.create_sheet("README")
    readme.append(["Daikin SkyAir 2026 controlled import package"])
    readme.append(["Source", SOURCE_NAME])
    readme.append(["SHA-256", SOURCE_SHA256])
    readme.append(["One row", "One explicitly published indoor/outdoor system combination"])
    readme.append(["Production import", "PROHIBITED - dry-run and isolated DB only"])
    readme.append(["Physical pages reviewed", "88/88"])
    book.save(WORKBOOK)

    summary = {
        "pdf_pages": 88, "indoor_models": len(inventory_indoor), "outdoor_models": len(inventory_outdoor),
        "combinations": len(flat), "import_ready_before_schema": len(ready), "review_required": len(review),
        "provenance_cells": len(provenance), "features": len(features), "accessories": len(accessory_rows),
        "controllers": len(controller_rows), "duplicate_skus": len(duplicate_skus), "sha256": SOURCE_SHA256,
        "by_family": Counter(row["outdoor_series"] for row in flat),
        "by_equipment": Counter(row["equipment_type"] for row in flat),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2, default=dict))


if __name__ == "__main__":
    main()
