"""Validate row decisions and, only when fully dispositioned, build the import workbook.

This script never approves rows. The operator workbook is the decision input;
technical values are always taken from the immutable source workbook.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
REVIEW = ROOT / "DAIKIN_SKYAIR_2026_OPERATOR_REVIEW.xlsx"
SOURCE = ROOT / "DAIKIN_SKYAIR_2026_IMPORT.xlsx"
MATRIX = ROOT / "docs/reports/final/artifacts/skyair_combination_matrix.csv"
OUTPUT = ROOT / "DAIKIN_SKYAIR_2026_PRODUCTION_IMPORT.xlsx"
ARTIFACTS = ROOT / "docs/reports/final/artifacts"
ALLOWED = {"REVIEW", "APPROVE", "REJECT"}
IMMUTABLE = ["sku", "indoor_model", "outdoor_model", "equipment_type", "capacity_kw", "technical_capacity_btu", "refrigerant", "phase"]


def rows_from_sheet(path: Path, sheet_name: str) -> list[dict[str, str]]:
    sheet = load_workbook(path, read_only=True, data_only=True)[sheet_name]
    values = list(sheet.values)
    headers = [str(value or "") for value in values[0]]
    return [dict(zip(headers, values)) for values in values[1:]]


def source_data_hash() -> str:
    return hashlib.sha256(MATRIX.read_bytes()).hexdigest().upper()


def matrix_rows() -> list[dict[str, str]]:
    with MATRIX.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--generate", action="store_true", help="generate only after REVIEW=0")
    args = parser.parse_args()

    review_rows = rows_from_sheet(REVIEW, "OPERATOR_REVIEW")
    source_rows = rows_from_sheet(SOURCE, "IMPORT_READY")
    source_by_sku = {row["sku"]: row for row in source_rows}
    matrix_by_sku = {row["sku"]: row for row in matrix_rows()}
    errors = []
    counts = {decision: 0 for decision in ALLOWED}

    for row in review_rows:
        decision = str(row.get("operator_decision") or "").strip().upper()
        counts[decision] = counts.get(decision, 0) + 1
        if decision not in ALLOWED:
            errors.append(f"{row.get('sku')}: invalid decision {decision!r}")
        if decision == "REJECT" and not str(row.get("operator_note") or "").strip():
            errors.append(f"{row.get('sku')}: REJECT requires operator_note")
        source = matrix_by_sku.get(row.get("sku"))
        if source is None:
            errors.append(f"{row.get('sku')}: SKU not found in immutable source workbook")
            continue
        for field in IMMUTABLE:
            if str(row.get(field) or "") != str(source.get(field) or ""):
                errors.append(f"{row.get('sku')}: immutable field changed: {field}")
        if row.get("schema_id") != f"skyair-{source.get('equipment_type')}-v1":
            errors.append(f"{row.get('sku')}: invalid category schema mapping")
        if row.get("source_status") != "VERIFIED" or row.get("pairing_status") != "VERIFIED" or row.get("category_status") != "VERIFIED" or row.get("schema_status") != "VERIFIED" or row.get("capacity_status") != "VERIFIED":
            errors.append(f"{row.get('sku')}: source gate is not VERIFIED")

    if len(review_rows) != len(source_rows):
        errors.append(f"row count mismatch: review={len(review_rows)} source={len(source_rows)}")
    if len({row.get("sku") for row in review_rows}) != len(review_rows):
        errors.append("duplicate SKU in operator workbook")
    if errors:
        print(json.dumps({"status": "BLOCKED", "errors": errors[:20], "error_count": len(errors), "counts": counts}, indent=2))
        raise SystemExit(2)
    if counts.get("REVIEW", 0) > 0:
        print(json.dumps({"status": "REVIEW_REQUIRED", "counts": counts, "source_data_hash": source_data_hash(), "production_workbook": "NOT_CREATED"}, indent=2))
        raise SystemExit(0)
    if not args.generate:
        print(json.dumps({"status": "READY_TO_GENERATE", "counts": counts, "source_data_hash": source_data_hash(), "production_workbook": "NOT_CREATED (use --generate after collision check)"}, indent=2))
        raise SystemExit(0)

    approved = [source_by_sku[row["sku"]] for row in review_rows if str(row.get("operator_decision")).upper() == "APPROVE"]
    book = Workbook()
    sheet = book.active
    sheet.title = "IMPORT_READY"
    headers = list(source_rows[0].keys())
    sheet.append(headers)
    for row in approved:
        sheet.append([row.get(header, "") for header in headers])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    metadata = book.create_sheet("README")
    metadata.append(["Daikin SkyAir 2026 approved production import workbook"])
    metadata.append(["Generated at", datetime.now(timezone.utc).isoformat()])
    metadata.append(["Rows", len(approved)])
    metadata.append(["Source data hash", source_data_hash()])
    metadata.append(["Operator workbook hash", hashlib.sha256(REVIEW.read_bytes()).hexdigest().upper()])
    book.save(OUTPUT)

    manifest_path = ARTIFACTS / "skyair_production_import_manifest.csv"
    with manifest_path.open("w", encoding="utf-8-sig", newline="") as handle:
        fields = ["sku", "model_code", "indoor_model", "outdoor_model", "category_id", "schema_id", "action", "approval", "technical_field_count", "source_data_hash"]
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in review_rows:
            decision = str(row["operator_decision"]).upper()
            writer.writerow({"sku": row["sku"], "model_code": source_by_sku[row["sku"]]["model_code"], "indoor_model": row["indoor_model"], "outdoor_model": row["outdoor_model"], "category_id": row.get("product_category_id", ""), "schema_id": row.get("schema_id", ""), "action": "IMPORT" if decision == "APPROVE" else "EXCLUDED_REJECTED", "approval": decision, "technical_field_count": sum(1 for key, value in source_by_sku[row["sku"]].items() if value not in (None, "")), "source_data_hash": source_data_hash()})
    print(json.dumps({"status": "GENERATED", "counts": counts, "rows": len(approved), "workbook": str(OUTPUT), "sha256": hashlib.sha256(OUTPUT.read_bytes()).hexdigest().upper()}, indent=2))


if __name__ == "__main__":
    main()
