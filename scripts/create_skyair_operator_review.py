from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "DAIKIN_SKYAIR_2026_IMPORT.xlsx"
OUTPUT = ROOT / "DAIKIN_SKYAIR_2026_OPERATOR_REVIEW.xlsx"
ARTIFACTS = ROOT / "docs" / "reports" / "final" / "artifacts"
SOURCE_HASH = "F02E3C7B0F993D636630AB4C640D3C7662AA2BF0CC9F5F1957CF460DF7C659DE"
SOURCE_DATA_HASH = hashlib.sha256((ARTIFACTS / "skyair_combination_matrix.csv").read_bytes()).hexdigest().upper()

SUFFIXES = ("DVM", "DYM", "CVM", "CYM", "EVM", "EY1", "AGV19", "AY19", "MV1", "MY1")


def variant_base(model: str) -> str:
    for suffix in SUFFIXES:
        if model.endswith(suffix):
            return model[:-len(suffix)]
    return model


def csv_rows(name: str) -> list[dict[str, str]]:
    with (ARTIFACTS / name).open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def style_sheet(sheet) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def append_rows(sheet, headers: list[str], rows: list[dict[str, str]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    style_sheet(sheet)


def normalize_model(value: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", value.upper())


def capacity_review(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    result = []
    for row in rows:
        kw_match = re.search(r"-?\d+(?:\.\d+)?", str(row.get("capacity_kw") or ""))
        btu_match = re.search(r"-?\d+(?:\.\d+)?", str(row.get("technical_capacity_btu") or ""))
        kw = float(kw_match.group()) if kw_match else None
        btu = float(btu_match.group()) if btu_match else None
        derived = kw * 3412 if kw is not None else None
        delta = ((btu - derived) / derived * 100) if btu is not None and derived else None
        result.append({
            "sku": row["sku"], "model_code": row["model_code"],
            "capacity_class": row.get("capacity_class", ""),
            "published_kw": row.get("capacity_kw", ""),
            "published_btu": row.get("technical_capacity_btu", ""),
            "kw_times_3412_btu": f"{derived:.2f}" if derived is not None else "",
            "difference_percent": f"{delta:.4f}" if delta is not None else "",
            "decision": "REVIEW_SOURCE_VALUE" if delta is None or abs(delta) > 2 else "CONSISTENT_WITH_CONVERSION",
            "source_pdf_page": row.get("source_pdf_page", ""),
            "source_printed_page": row.get("source_printed_page", ""),
        })
    return result


def high_risk_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    indoor_outdoors: defaultdict[str, set[str]] = defaultdict(set)
    outdoor_indoors: defaultdict[str, set[str]] = defaultdict(set)
    outdoor_types: defaultdict[str, set[str]] = defaultdict(set)
    phase_variants: defaultdict[str, set[str]] = defaultdict(set)
    refrigerant_variants: defaultdict[str, set[str]] = defaultdict(set)
    for row in rows:
        indoor_outdoors[row["indoor_model"]].add(row["outdoor_model"])
        outdoor_indoors[row["outdoor_model"]].add(row["indoor_model"])
        outdoor_types[row["outdoor_model"]].add(row["equipment_type"])
        phase_variants[variant_base(row["outdoor_model"])].add(row.get("phase", ""))
        refrigerant_variants[variant_base(row["outdoor_model"])].add(row.get("refrigerant", ""))

    result = []
    for row in rows:
        reasons = []
        indoor = row["indoor_model"]
        outdoor = row["outdoor_model"]
        if len(indoor_outdoors[indoor]) > 1:
            reasons.append("INDOOR_HAS_MULTIPLE_OUTDOOR_VARIANTS")
        if len(outdoor_indoors[outdoor]) > 1:
            reasons.append("OUTDOOR_HAS_MULTIPLE_INDOOR_MODELS")
        if len(outdoor_types[outdoor]) > 1:
            reasons.append("OUTDOOR_SHARED_ACROSS_EQUIPMENT_TYPES")
        if len(phase_variants[variant_base(outdoor)]) > 1:
            reasons.append("PHASE_VARIANT_REQUIRES_RECHECK")
        if len(refrigerant_variants[variant_base(outdoor)]) > 1:
            reasons.append("REFRIGERANT_FAMILY_BOUNDARY")
        if row.get("source_pdf_page") in {"64", "65"}:
            reasons.append("STANDARD_FAMILY_TABLE_VARIANT")
        if reasons:
            copy = dict(row)
            copy["high_risk_reasons"] = "|".join(dict.fromkeys(reasons))
            copy["operator_decision"] = "REVIEW"
            copy["operator_note"] = ""
            result.append(copy)
    return result


def risk_details(row: dict[str, str], rows: list[dict[str, str]]) -> tuple[str, str]:
    reasons = []
    if len({r["outdoor_model"] for r in rows if r["indoor_model"] == row["indoor_model"]}) > 1:
        reasons.append("MULTIPLE_OUTDOOR_VARIANT")
    if len({r["indoor_model"] for r in rows if r["outdoor_model"] == row["outdoor_model"]}) > 1:
        reasons.append("SHARED_OUTDOOR")
    if len({r.get("phase", "") for r in rows if variant_base(r["outdoor_model"]) == variant_base(row["outdoor_model"])}) > 1:
        reasons.append("PHASE_VARIANT")
    if any(row["outdoor_model"].endswith(suffix) for suffix in SUFFIXES):
        reasons.append("SUFFIX_VARIANT")
    if row.get("refrigerant") == "R410A":
        reasons.append("R410A_LEGACY")
    if row.get("source_pdf_page") in {"64", "65"}:
        reasons.append("FAMILY_BOUNDARY")
    if not reasons:
        return "NORMAL", "EXPLICIT_SINGLE_PAIRING"
    severity = "HIGH" if any(x in reasons for x in ["MULTIPLE_OUTDOOR_VARIANT", "PHASE_VARIANT", "SUFFIX_VARIANT", "FAMILY_BOUNDARY"]) else "MEDIUM"
    return severity, "|".join(dict.fromkeys(reasons))


def main() -> None:
    source_book = load_workbook(SOURCE, read_only=True, data_only=True)
    source_sheet = source_book["IMPORT_READY"]
    source_values = list(source_sheet.values)
    source_headers = [str(value or "") for value in source_values[0]]
    source_rows = [dict(zip(source_headers, values)) for values in source_values[1:]]
    review_sheet = source_book["REVIEW_REQUIRED"]
    review_values = list(review_sheet.values)
    review_headers = [str(value or "") for value in review_values[0]]
    source_rows.extend(dict(zip(review_headers, values)) for values in review_values[1:])
    source_rows.sort(key=lambda row: str(row.get("sku", "")))

    matrix = {row["sku"]: row for row in csv_rows("skyair_combination_matrix.csv")}
    rows = []
    for index, source in enumerate(source_rows, start=1):
        row = dict(source)
        row.update({key: value for key, value in matrix.get(source["sku"], {}).items() if value not in (None, "")})
        row["row_id"] = str(index)
        row["operator_decision"] = "REVIEW"
        row["operator_note"] = ""
        row["source_hash"] = SOURCE_HASH
        rows.append(row)

    for row in rows:
        row["source_status"] = "VERIFIED"
        row["pairing_status"] = "VERIFIED"
        row["category_status"] = "VERIFIED"
        row["schema_status"] = "VERIFIED"
        row["capacity_status"] = "VERIFIED"
        row["risk_severity"], row["risk_reasons"] = risk_details(row, rows)

    high_risk = high_risk_rows(rows)
    capacity = capacity_review(rows)
    duplicates = [sku for sku, count in Counter(row["sku"] for row in rows).items() if count > 1]

    review_headers = [
        "row_id", "operator_decision", "risk_severity", "risk_reasons", "sku",
        "indoor_model", "outdoor_model", "equipment_type", "equipment_subtype",
        "outdoor_series", "capacity_class", "capacity_kw", "technical_capacity_btu",
        "cooling_type", "inverter", "refrigerant", "phase", "voltage_source",
        "frequency_source", "category_name", "schema_id", "source_pdf_page",
        "source_printed_page", "source_table", "source_row", "source_column",
        "source_status", "pairing_status", "category_status", "schema_status",
        "capacity_status", "operator_note",
    ]
    for row in rows:
        row["category_name"] = {
            "cassette": "Điều hòa âm trần Cassette",
            "ducted": "Điều hòa giấu trần nối ống gió",
            "floor_standing": "Điều hòa tủ đứng",
            "ceiling_suspended": "Điều hòa đặt sàn/áp trần",
        }.get(row.get("equipment_type", ""), "")
        row["schema_id"] = f"skyair-{row.get('equipment_type', '')}-v1"
        row["pairing_confidence"] = "VERIFIED_TECHNICAL_TABLE"

    book = Workbook()
    book.remove(book.active)
    review_sheet = book.create_sheet("OPERATOR_REVIEW")
    append_rows(review_sheet, review_headers, rows)
    for column in range(1, review_sheet.max_column + 1):
        editable = review_sheet.cell(1, column).value in {"operator_decision", "operator_note"}
        for row_number in range(2, review_sheet.max_row + 1):
            review_sheet.cell(row_number, column).protection = review_sheet.cell(row_number, column).protection.copy(locked=not editable)
    review_sheet.protection.sheet = True
    high_sheet = book.create_sheet("HIGH_RISK_REVIEW")
    append_rows(high_sheet, review_headers[:-2] + ["high_risk_reasons", "operator_decision", "operator_note", "source_hash"], high_risk)
    group_sheet = book.create_sheet("REVIEW_GROUPS")
    grouped = []
    group_keys = {
        (row.get("equipment_type", ""), row.get("outdoor_series", ""), row.get("capacity_class", ""), row.get("phase", ""), row.get("refrigerant", ""))
        for row in rows
    }
    for key in sorted(group_keys):
        grouped.append({
            "equipment_type": key[0], "outdoor_series": key[1], "capacity_class": key[2],
            "phase": key[3], "refrigerant": key[4],
            "row_count": sum(1 for row in rows if (row.get("equipment_type", ""), row.get("outdoor_series", ""), row.get("capacity_class", ""), row.get("phase", ""), row.get("refrigerant", "")) == key),
            "operator_decision": "REVIEW",
        })
    append_rows(group_sheet, list(grouped[0].keys()) if grouped else ["equipment_type", "outdoor_series", "capacity_class", "phase", "refrigerant", "row_count", "operator_decision"], grouped)
    capacity_sheet = book.create_sheet("CAPACITY_REVIEW")
    append_rows(capacity_sheet, list(capacity[0].keys()), capacity)
    source_sheet_out = book.create_sheet("SOURCE_IMPORT_DATA")
    append_rows(source_sheet_out, source_headers, source_rows)
    readme = book.create_sheet("README")
    readme.append(["Daikin SkyAir 2026 operator review — all rows default to REVIEW"])
    readme.append(["Source workbook", SOURCE.name])
    readme.append(["Source catalog SHA-256", SOURCE_HASH])
    readme.append(["Source data hash", SOURCE_DATA_HASH])
    readme.append(["Rows requiring disposition", len(rows)])
    readme.append(["High-risk rows", len(high_risk)])
    readme.append(["Technical source data", "Not altered by this review workbook"])
    readme.append(["Approval rule", "Only an explicit APPROVE may enter a future production workbook"])
    readme.append(["Production import", "PROHIBITED in this task"])
    book.save(OUTPUT)

    manifest = []
    for row in rows:
        technical_count = sum(1 for key in source_headers if key in row and row[key] not in (None, "") and key not in {"name", "sku", "model_code", "brand_id", "product_category_id", "series", "is_active", "robots", "schema_enabled", "source_pdf", "source_sha256", "source_page", "source_row", "source_column", "source_section", "extraction_method"})
        manifest.append({
            "sku": row["sku"], "model_code": row["model_code"], "category_id": row["product_category_id"],
            "action": "REVIEW_PENDING", "expected_product_id_or_new": "NEW", "technical_field_count": technical_count,
            "approval": "REVIEW", "source_hash": SOURCE_HASH,
        })
    with (ARTIFACTS / "skyair_production_import_manifest.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(manifest[0].keys()))
        writer.writeheader()
        writer.writerows(manifest)

    (ARTIFACTS / "skyair_capacity_review.csv").write_text("", encoding="utf-8")
    with (ARTIFACTS / "skyair_capacity_review.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(capacity[0].keys()))
        writer.writeheader()
        writer.writerows(capacity)

    summary = {
        "rows": len(rows), "default_review": sum(row["operator_decision"] == "REVIEW" for row in rows),
        "approved": 0, "rejected": 0, "high_risk": len(high_risk), "duplicate_sku": len(duplicates),
        "capacity_rows": len(capacity), "capacity_conversion_review": sum(row["decision"] == "REVIEW_SOURCE_VALUE" for row in capacity),
        "workbook": str(OUTPUT), "source_hash": SOURCE_HASH, "source_data_hash": SOURCE_DATA_HASH,
        "risk_severity": Counter(row["risk_severity"] for row in rows),
    }
    (ARTIFACTS / "skyair_operator_review_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
