"""Extract source-backed SkyAir bundle component assignments from an XLSX.

This does not choose a controller or panel.  A Product pair is marked exact
only when every source bundle row for that pair names the same component.
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


PAIR = re.compile(
    r"\b(F[A-Z0-9()\-]+)\s*/\s*((?:RZF|RZA|RZFC|RNQ|RCN|RC|RN)[A-Z0-9()\-]+)"
    r"(?:\s*\+\s*((?:BRC|ARC)[A-Z0-9()\-]+))?"
    r"(?:\s*\+\s*((?:BYCQ|BYFQ)[A-Z0-9()\-]+))?",
    re.IGNORECASE,
)


def values(rows: list[dict[str, str]], key: str) -> list[str]:
    return sorted({row[key] for row in rows if row[key]})


def classification(options: list[str]) -> str:
    if len(options) == 1:
        return "EXACT_BUNDLE_ASSIGNMENT"
    if len(options) > 1:
        return "COMPATIBILITY_ONLY"
    return "NO_SOURCE_EVIDENCE"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", required=True, type=Path)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("docs/reports/final/artifacts/skyair_bundle_component_matrix.csv"),
    )
    args = parser.parse_args()

    book = load_workbook(args.source, data_only=True, read_only=True)
    by_pair: dict[str, list[dict[str, str]]] = defaultdict(list)
    for sheet in book.worksheets:
        for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            text = " ".join(str(value).strip() for value in row if value is not None)
            for match in PAIR.finditer(text):
                indoor, outdoor, remote, panel = (value.upper() if value else "" for value in match.groups())
                pair = f"{indoor}/{outdoor}"
                by_pair[pair].append(
                    {
                        "sheet": sheet.title,
                        "source_row": str(index),
                        "model_code": pair,
                        "remote_model": remote,
                        "panel_model": panel,
                        "source_text": text,
                    }
                )

    output: list[dict[str, str]] = []
    for pair, rows in sorted(by_pair.items()):
        remote_options = values(rows, "remote_model")
        panel_options = values(rows, "panel_model")
        for row in rows:
            output.append(
                {
                    **row,
                    "remote_options": "|".join(remote_options),
                    "remote_classification": classification(remote_options),
                    "panel_options": "|".join(panel_options),
                    "panel_classification": classification(panel_options),
                    "row_classification": "BUNDLE_ASSIGNMENT",
                }
            )

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(output[0]))
        writer.writeheader()
        writer.writerows(output)


if __name__ == "__main__":
    main()
